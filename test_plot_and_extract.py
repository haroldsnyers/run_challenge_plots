import generate_graphs
from generate_graphs import GenerateGraphs
import openpyxl

excel_document = openpyxl.load_workbook('Loutres.xlsx')

sheet_names = excel_document.sheetnames
print(sheet_names)

pf_sheet = excel_document[sheet_names[1]]

print(pf_sheet.cell(row=2, column=20).value)

list_names = [pf_sheet.cell(row=1, column=x).value for x in range(4, 28)]
list_names_columns = [x for x in range(4, 28)]

dict_names = dict(zip(list_names, list_names_columns))


def retrieve_values(sheet, name, x):
    return sheet.cell(row=x, column=dict_names[name]).value if sheet.cell(row=x, column=dict_names[
        name]).value is not None else 0


start, end = 2, 59
matthieu_values = [
    [retrieve_values(pf_sheet, 'Mathieu', x) for x in range(start, end, 8)],
    [retrieve_values(pf_sheet, 'Mathieu', x) for x in range(start + 2, end + 2, 8)],
    [retrieve_values(pf_sheet, 'Mathieu', x) for x in range(start + 4, end + 4, 8)],
    [retrieve_values(pf_sheet, 'Mathieu', x) for x in range(start + 6, end + 6, 8)],
]
print(matthieu_values)

test = [
    [retrieve_values(pf_sheet, 'Mathieu', x),
     retrieve_values(pf_sheet, 'Mathieu', x + 2),
     retrieve_values(pf_sheet, 'Mathieu', x + 4),
     retrieve_values(pf_sheet, 'Mathieu', x + 6)] for x in range(start, end, 8)]

print(test)



# bar chart
generator = GenerateGraphs('Mathieu', matthieu_values)

generator.generate_bar_chart_km()

